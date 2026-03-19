#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
TXT <-> XLSX 转换器（UTF-8 CRLF TXT 与 标准 Excel 互转）
修复 game 字段丢失问题，正确处理多行描述及空行。
新增功能：无参数运行时生成模板文件。
"""

import sys
import argparse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

FIELDS = ['game', 'file', 'sort-by', 'developer', 'description']

def parse_txt(txt_path):
    """
    解析 UTF-8 TXT，返回游戏字典列表。
    正确处理多行 description 和内部空行。
    """
    games = []
    cur_game = None          # 当前正在构建的游戏字典
    cur_field = None         # 当前正在处理的字段名
    cur_lines = []           # 当前字段累积的行（用于多行）

    def save_field():
        """将 cur_lines 中的多行合并，存入当前游戏的当前字段"""
        if cur_game is not None and cur_field is not None:
            # 合并多行，保留内部换行
            value = '\n'.join(cur_lines)
            cur_game[cur_field] = value

    with open(txt_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.rstrip('\r\n')   # 只去除换行符，保留行内空格

            # 处理空行：如果在某个字段内，空行应作为字段内容的一部分
            if line == '':
                if cur_field is not None:
                    cur_lines.append('')
                continue

            # 检测是否为字段开始行（如 "game: 重装机兵"）
            field_start = None
            for field in FIELDS:
                if line.lower().startswith(field.lower() + ':'):
                    field_start = field
                    break

            if field_start is not None:
                # 先保存上一个字段（如果有）
                save_field()

                # 提取当前字段的值部分（去掉 "字段名:" 及前导空格）
                value_part = line[len(field_start)+1:].lstrip()

                if field_start == 'game':
                    # 新游戏开始
                    if cur_game is not None:
                        games.append(cur_game)   # 将上一个游戏存入列表
                    cur_game = {}
                    cur_field = field_start
                    # game 字段只有一行，但其值放入 cur_lines 等待后续保存
                    cur_lines = [value_part] if value_part else ['']
                else:
                    # 非 game 字段，属于当前游戏
                    if cur_game is None:
                        # 如果还没遇到 game 字段，跳过（理论上不应发生）
                        continue
                    cur_field = field_start
                    # 开始累积该字段的值，第一行已存在
                    cur_lines = [value_part] if value_part else ['']
            else:
                # 普通行：作为当前字段的续行
                if cur_field is not None:
                    cur_lines.append(line)
                # 否则忽略（文件开头的无关行）

    # 循环结束，保存最后一个字段
    save_field()
    # 将最后一个游戏加入列表
    if cur_game is not None:
        games.append(cur_game)

    return games

def write_xlsx(games, xlsx_path):
    """将游戏列表写入 XLSX，description 列启用自动换行"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Games"

    # 表头
    for col, field in enumerate(FIELDS, 1):
        ws.cell(row=1, column=col, value=field)

    # 数据
    for row, game in enumerate(games, 2):
        for col, field in enumerate(FIELDS, 1):
            cell = ws.cell(row=row, column=col, value=game.get(field, ''))
            if field == 'description':
                cell.alignment = Alignment(wrapText=True)

    # 简单调整列宽
    for col in range(1, len(FIELDS)+1):
        ws.column_dimensions[chr(64+col)].width = 25

    wb.save(xlsx_path)

def parse_xlsx(xlsx_path):
    """从 XLSX 读取游戏列表（第一行为表头）"""
    wb = load_workbook(xlsx_path)
    ws = wb.active

    games = []
    for row in range(2, ws.max_row + 1):
        game = {}
        for col, field in enumerate(FIELDS, 1):
            val = ws.cell(row=row, column=col).value
            if val is None:
                game[field] = ''
            else:
                # 转换为字符串，并去除尾随的换行符和空白字符
                s = str(val)
                # 去除尾随的空白字符（包括换行符）
                s = s.rstrip()
                game[field] = s
        games.append(game)
    return games

def write_txt(games, txt_path):
    """将游戏列表写入 UTF-8 TXT，CRLF 换行"""
    with open(txt_path, 'w', encoding='utf-8', newline='') as f:
        for i, game in enumerate(games):
            for field in FIELDS:
                value = game.get(field, '')
                if value is None:
                    value = ''
                else:
                    value = str(value)
                    # 确保值没有尾随的换行符
                    value = value.rstrip('\r\n')
                f.write(f"{field}: {value}\r\n")
                # 如果是description字段，添加一个空行
                if field == 'description':
                    f.write("\r\n")
            # 游戏之间不加空行（如需添加，取消下一行注释）
            # if i < len(games) - 1:
            #     f.write("\r\n")

def txt_to_xlsx(txt_path, xlsx_path):
    games = parse_txt(txt_path)
    if not games:
        print("警告：解析后没有游戏数据，请检查 TXT 格式。")
    write_xlsx(games, xlsx_path)
    print(f"成功转换：{txt_path} -> {xlsx_path}")

def xlsx_to_txt(xlsx_path, txt_path):
    games = parse_xlsx(xlsx_path)
    write_txt(games, txt_path)
    print(f"成功转换：{xlsx_path} -> {txt_path}")

def generate_templates():
    """生成模板文件：template_games.txt 和 template_games.xlsx"""
    # 创建示例数据
    template_games = [
        {
            'game': '重装机兵',
            'file': '重装机兵.zip',
            'sort-by': '001',
            'developer': 'Crea-Tech',
            'description': '《重装机兵》是一款末世科幻风格的RPG游戏，由宫冈宽制作。'
        },
        {
            'game': '赤影战士',
            'file': '赤影战士.zip',
            'sort-by': '002',
            'developer': 'NATSUME',
            'description': '《赤影战士》是NATSUME公司于1990年推出的清版过关游戏。'
        }
    ]
    
    # 生成TXT模板
    txt_path = 'template_games.txt'
    write_txt(template_games, txt_path)
    print(f"已生成TXT模板：{txt_path}")
    
    # 生成XLSX模板
    xlsx_path = 'template_games.xlsx'
    write_xlsx(template_games, xlsx_path)
    print(f"已生成XLSX模板：{xlsx_path}")
    
    print("\n模板文件已生成在当前目录，可用于测试转换功能。")
    print("使用方法示例：")
    print(f"  1. TXT转XLSX: python {sys.argv[0]} txt2xlsx 要转换的TXT文件的名字.txt 生成的表格的名字.xlsx")
    print(f"  2. XLSX转TXT: python {sys.argv[0]} xlsx2txt 要转换的XLSX文件的名字.xlsx 生成的TXT文件的名字.txt")

def main():
    parser = argparse.ArgumentParser(description='TXT (UTF-8 CRLF) 与 XLSX 互转工具')
    parser.add_argument('mode', choices=['txt2xlsx', 'xlsx2txt'], help='转换模式')
    parser.add_argument('input', help='输入文件路径')
    parser.add_argument('output', help='输出文件路径')
    args = parser.parse_args()

    if args.mode == 'txt2xlsx':
        txt_to_xlsx(args.input, args.output)
    else:
        xlsx_to_txt(args.input, args.output)

if __name__ == '__main__':
    if len(sys.argv) == 1:
        # 无参数时生成模板文件
        generate_templates()
    else:
        main()
